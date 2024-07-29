# Copyright (c) 2024 ot2i7ba
# https://github.com/ot2i7ba/
# This code is licensed under the MIT License (see LICENSE for details).

"""
This script generates random fake work schedules and timesheets.
"""

import csv
import random
from datetime import datetime, timedelta
import calendar
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image
import os
import holidays
import win32com.client  # For PDF conversion on Windows
import pythoncom  # For COM initialization in threads
import time

# Constants
MAX_ATTEMPTS = 1000
DEFAULT_MAX_WORKDAYS = 8
DEFAULT_MIN_WORKDAYS = 6
DEFAULT_START_TIME = "17:00"
DEFAULT_END_TIME = "22:00"
DEFAULT_MAX_HOURS = 17.0
DEFAULT_MIN_HOURS = 15.0
ALLOWED_MINUTES = [0, 15, 30, 45]
TIME_FORMAT = "%H:%M"
DATE_FORMAT = "%d.%m.%Y"

def clear_terminal():
    os.system('cls' if os.name == 'nt' else 'clear')

def get_all_dates(year, month):
    days_in_month = calendar.monthrange(year, month)[1]
    return [datetime(year, month, day) for day in range(1, days_in_month + 1)]

def get_holidays(year):
    german_holidays = holidays.Germany(state='NW', years=year)
    special_days = [datetime(year, 12, 24), datetime(year, 12, 31), datetime(year + 1, 1, 1)]
    return german_holidays, special_days

def get_random_work_dates(all_dates, min_workdays_count, max_workdays_count, holidays, special_days):
    workdays_count = random.randint(min_workdays_count, max_workdays_count)
    work_dates = set()
    while len(work_dates) < workdays_count:
        date = random.choice(all_dates)
        if date.weekday() < 5 and date not in holidays and date not in special_days:  # Monday to Friday, not a holiday or special day
            work_dates.add(date)
    return sorted(work_dates)

def generate_time_sheet(year, month, min_workdays_count, max_workdays_count, start_time, end_time, max_hours, min_hours):
    all_dates = get_all_dates(year, month)
    holidays_in_month, special_days_in_month = get_holidays(year)
    total_hours = 0
    rows = []

    start_hour, start_minute = map(int, start_time.split(':'))
    end_hour, end_minute = map(int, end_time.split(':'))

    work_dates = get_random_work_dates(all_dates, min_workdays_count, max_workdays_count, holidays_in_month, special_days_in_month)

    attempts = 0

    while total_hours < min_hours and attempts < MAX_ATTEMPTS:
        rows = []
        total_hours = 0
        attempts += 1

        for date in all_dates:
            if date in work_dates:
                random_start, random_end, daily_hours = generate_random_work_hours(year, month, date.day, start_hour, start_minute, end_hour, end_minute)
                if total_hours + daily_hours > max_hours:
                    continue

                total_hours += daily_hours
                rows.append(format_workday(date, random_start, random_end, daily_hours))
            else:
                rows.append([date.strftime(DATE_FORMAT), "00:00", "00:00", "00:00"])

    if attempts == MAX_ATTEMPTS:
        return None, None

    return rows, total_hours

def generate_random_work_hours(year, month, day, start_hour, start_minute, end_hour, end_minute):
    random_start_hour = random.randint(start_hour, end_hour - 1)
    random_start_minute = random.choice(ALLOWED_MINUTES)
    random_start = datetime(year, month, day, random_start_hour, random_start_minute)

    max_end = datetime(year, month, day, end_hour, end_minute)
    random_end = random_start + timedelta(minutes=random.choice(range(15, 180 + 1, 15)))
    if random_end > max_end:
        random_end = max_end

    daily_seconds = (random_end - random_start).seconds
    daily_hours = daily_seconds / 3600

    return random_start, random_end, daily_hours

def format_workday(date, start, end, daily_hours):
    hours, minutes = divmod(int(daily_hours * 60), 60)
    return [date.strftime(DATE_FORMAT), start.strftime(TIME_FORMAT), end.strftime(TIME_FORMAT), f"{hours:02}:{minutes:02}"]

def save_to_excel(filename, rows, total_hours, sheet_name, signature, signature_image_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    add_title(sheet, sheet_name)
    sheet.append([])  # Add an empty row after the title
    headers = ['DATUM', 'ANFANG', 'ENDE', 'GESAMT']
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    sheet.append([])  # Add an empty row
    add_total_hours(sheet, total_hours, len(rows))
    add_signature(sheet, signature, rows, signature_image_path)

    apply_styles(sheet, headers, len(rows))
    workbook.save(filename)

def add_title(sheet, sheet_name):
    title = f"Stundenzettel {sheet_name}"
    sheet.merge_cells('A1:D1')
    title_cell = sheet['A1']
    title_cell.value = title
    title_cell.font = Font(name='Arial', size=20, bold=True)
    title_cell.alignment = Alignment(horizontal='left', vertical='center')

def add_total_hours(sheet, total_hours, rows_count):
    total_row = ["", "", "Gesamt:", f"{int(total_hours):02}:{int((total_hours * 60) % 60):02}"]
    sheet.append(total_row)

def add_signature(sheet, signature, rows, signature_image_path):
    signature_row_index = len(rows) + 10
    last_work_date = rows[-1][0]
    signature_text = f"{last_work_date}, {signature}"
    signature_cell = sheet.cell(row=signature_row_index, column=1, value=signature_text)

    if signature_image_path:
        img = Image(signature_image_path)
        scale_image_to_fit(img, max_width=200, max_height=80)
        sheet.add_image(img, f'A{signature_row_index - 2}')
    signature_cell.font = Font(name='Arial', size=12, bold=True)

def scale_image_to_fit(img, max_width, max_height):
    original_width, original_height = img.width, img.height
    aspect_ratio = original_width / original_height

    if original_width > original_height:
        img.width = max_width
        img.height = max_width / aspect_ratio
    else:
        img.height = max_height
        img.width = max_height * aspect_ratio

def apply_styles(sheet, headers, rows_count):
    header_font = Font(name='Arial', size=12, bold=True)
    total_font = Font(name='Arial', size=12, bold=True)
    workday_font = Font(name='Arial', size=12, bold=True)
    normal_font = Font(name='Arial', size=12)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="B0C4DE", end_color="B0C4DE", fill_type="solid")
    fill_even = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Apply styles to header
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=3, column=col)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='left' if col < 4 else 'right', vertical='center')
        cell.border = thin_border
        cell.fill = header_fill
        sheet.column_dimensions[get_column_letter(col)].width = 15

    # Apply styles to total row
    for col in range(3, 5):
        cell = sheet.cell(row=rows_count + 5, column=col)
        cell.font = total_font
        cell.alignment = Alignment(horizontal='right' if col == 4 else 'left', vertical='center')
        cell.border = thin_border

    # Apply styles to data rows
    for row in sheet.iter_rows(min_row=4, max_row=rows_count + 3, min_col=1, max_col=4):
        is_workday = row[1].value != "00:00"
        for cell in row:
            cell.font = workday_font if is_workday else normal_font
            cell.border = thin_border
            if cell.column < 4:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='right', vertical='center')

            cell.fill = fill_even if cell.row % 2 == 0 else fill_odd

def convert_excel_to_pdf(excel_path, pdf_path):
    pythoncom.CoInitialize()
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        wb = excel.Workbooks.Open(os.path.abspath(excel_path))
        ws = wb.Worksheets(1)
        ws.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
        wb.Close(False)
        excel.Quit()
    finally:
        pythoncom.CoUninitialize()

def process_month(year, month, min_workdays_count, max_workdays_count, start_time, end_time, max_hours, min_hours, show_preview, signature, signature_image_path, directory, overwrite):
    retries = 3
    for attempt in range(retries):
        rows, total_hours = generate_time_sheet(year, month, min_workdays_count, max_workdays_count, start_time, end_time, max_hours, min_hours)
        if rows is not None and total_hours is not None:
            break
        print(f"Retry {attempt + 1} for {year}-{month:02}...")

    if rows is None or total_hours is None:
        print(f"Unable to generate a valid time sheet for {year}-{month:02} after {retries} retries. Skipping month.")
        return False

    if show_preview:
        print(f"\nGenerated time sheet for {year}-{month:02}:")
        print(f"{'DATE':<12} {'START':<8} {'END':<8} {'TOTAL':<6}")
        for row in rows:
            print(f"{row[0]:<12} {row[1]:<8} {row[2]:<8} {row[3]:<6}")
        total_minutes = total_hours * 60
        total_hours, total_minutes = divmod(total_minutes, 60)
        print(f"\nTotal hours in the month: {int(total_hours):02}:{int(total_minutes):02}")

    filename_base = f'Stundenzettel_{year}_{month:02}'
    csv_filename = os.path.join(directory, f'{filename_base}.csv')
    xlsx_filename = os.path.join(directory, f'{filename_base}.xlsx')
    pdf_filename = os.path.join(directory, f'{filename_base}.pdf')
    sheet_name = f'{year}.{month:02}'

    if not overwrite and any(os.path.exists(f) for f in [csv_filename, xlsx_filename, pdf_filename]):
        print(f"Files for {month:02} {year} already exist. Skipping creation.")
        return False

    with open(csv_filename, mode='w', newline='') as file:
        writer = csv.writer(file, delimiter=';')
        writer.writerow(['DATE', 'START', 'END', 'TOTAL'])
        writer.writerows(rows)
    print(f"\nCSV file '{csv_filename}' was successfully created.")

    save_to_excel(xlsx_filename, rows, total_hours, sheet_name, signature, signature_image_path)
    print(f"\nXLSX file '{xlsx_filename}' was successfully created.")

    return xlsx_filename, pdf_filename

def process_pdf_conversions(tasks):
    for xlsx_filename, pdf_filename in tasks:
        try:
            convert_excel_to_pdf(xlsx_filename, pdf_filename)
            print(f"\nPDF file '{pdf_filename}' was successfully created.")
        except Exception as e:
            print(f"\nError creating PDF file '{pdf_filename}': {e}")

def validate_input(prompt, default, min_value=None, max_value=None, is_float=False):
    while True:
        user_input = input(prompt) or default
        try:
            if is_float:
                value = float(user_input)
            else:
                value = int(user_input)
            if (min_value is not None and value < min_value) or (max_value is not None and value > max_value):
                print(f"Please enter a value between {min_value} and {max_value}.")
                continue
            return value
        except ValueError:
            print("Invalid input. Please try again.")

def create_directory(directory):
    if not os.path.exists(directory):
        os.makedirs(directory)
    return True

def check_existing_files(directory, year):
    for month in range(1, 13):
        filename_base = f'Stundenzettel_{year}_{month:02}'
        if any(os.path.exists(os.path.join(directory, f'{filename_base}.{ext}')) for ext in ['csv', 'xlsx', 'pdf']):
            return True
    return False

def main():
    clear_terminal()
    
    bulk_mode = input("Do you want to perform bulk processing? (yes/no) [no]: ") or 'no'
    is_bulk = bulk_mode.lower() == 'yes'
    
    current_year = datetime.now().year
    current_month = datetime.now().month

    year = validate_input(f"Enter the year [{current_year}]: ", current_year, min_value=2000, max_value=current_year + 10)

    if not is_bulk:
        month = validate_input(f"Enter the month (1-12) [{current_month}]: ", current_month, min_value=1, max_value=12)

    max_workdays_count = validate_input("Enter the maximum workdays in the month [8]: ", DEFAULT_MAX_WORKDAYS, min_value=1, max_value=31)
    min_workdays_count = validate_input("Enter the minimum workdays in the month [6]: ", DEFAULT_MIN_WORKDAYS, min_value=1, max_value=max_workdays_count)

    start_time = input(f"Enter the earliest start time (HH:MM) [{DEFAULT_START_TIME}]: ") or DEFAULT_START_TIME
    end_time = input(f"Enter the latest end time (HH:MM) [{DEFAULT_END_TIME}]: ") or DEFAULT_END_TIME

    max_hours = validate_input("Enter the maximum total hours in the month [17]: ", DEFAULT_MAX_HOURS, min_value=0, is_float=True)
    min_hours = validate_input("Enter the minimum total hours in the month [15]: ", DEFAULT_MIN_HOURS, min_value=0, max_value=max_hours, is_float=True)

    while True:
        signature = input("Enter the name for the signature: ")
        if signature:
            break

    signature_image_path = None
    png_files = [f for f in os.listdir() if f.endswith('.png')]
    if png_files:
        print("The following PNG files were found:")
        for i, file in enumerate(png_files, 1):
            print(f"{i}. {file}")
        file_choice = input(f"Enter the number of the image to use for the signature (1-{len(png_files)}), or press Enter to skip: ")
        if file_choice.isdigit() and 1 <= int(file_choice) <= len(png_files):
            signature_image_path = png_files[int(file_choice) - 1]

    directory = f"Stundenzettel {year}"
    if not create_directory(directory):
        return

    overwrite = False
    if check_existing_files(directory, year):
        overwrite = input("Files already exist in the directory. Do you want to overwrite existing files? (yes/no) [yes]: ") or 'yes'
        overwrite = overwrite.lower() == 'yes'
        if not overwrite:
            print("Please backup your existing files and try again.")
            time.sleep(3)
            main()
            return

    failed_months = []
    pdf_conversion_tasks = []
    if is_bulk:
        show_preview_input = input("Do you want to preview the tables in the terminal? (yes/no) [no]: ") or 'no'
        show_preview = show_preview_input.lower() == 'yes'

        for month in range(1, 13):
            result = process_month(year, month, min_workdays_count, max_workdays_count, start_time, end_time, max_hours, min_hours, show_preview, signature, signature_image_path, directory, overwrite)
            if result:
                pdf_conversion_tasks.append(result)
            else:
                failed_months.append(f"{year}.{month:02}")

        process_pdf_conversions(pdf_conversion_tasks)

        if failed_months:
            print("\nCan't create documents for:")
            for failed_month in failed_months:
                print(failed_month)
    else:
        result = process_month(year, month, min_workdays_count, max_workdays_count, start_time, end_time, max_hours, min_hours, show_preview=True, signature=signature, signature_image_path=signature_image_path, directory=directory, overwrite=overwrite)
        if result:
            convert_excel_to_pdf(*result)
            print(f"\nPDF file '{result[1]}' was successfully created.")

if __name__ == "__main__":
    main()
